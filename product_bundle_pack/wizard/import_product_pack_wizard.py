# -*- coding: utf-8 -*-
# wizard/import_product_pack_wizard.py
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64, io, csv

try:
    import openpyxl
except Exception:
    openpyxl = None


class ImportProductPackWizard(models.TransientModel):
    _name = 'import.product.pack.wizard'
    _description = 'Import Product Pack Components'

    # --- Fields yang DIHARUSKAN oleh view XML ---
    import_type = fields.Selection(
        [('csv', 'CSV'), ('xlsx', 'Excel (.xlsx)')],
        string='Import Type',
        default='csv'
    )
    file_data = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='Filename')
    update_existing = fields.Boolean(string='Update Existing Components')
    replace_all = fields.Boolean(string='Replace All Components')

    allow_parent_only = fields.Boolean(
        string='Allow Import Parent Without Components',
        default=True
    )

    # --- Helper casters aman ---
    @staticmethod
    def S(v):
        if v is None:
            return ''
        return v.strip() if isinstance(v, str) else str(v).strip()

    @staticmethod
    def F(v):
        if v is None or v == '':
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(',', '')
        try:
            return float(s)
        except Exception:
            return 0.0

    @staticmethod
    def I(v):
        return int(ImportProductPackWizard.F(v))

    @staticmethod
    def B(v):
        if isinstance(v, bool):
            return v
        s = ImportProductPackWizard.S(v).lower()
        return s in ('1', 'true', 'yes', 'y', 't')

    # --- Tombol Download Template: lempar ke controller via act_url ---
    def download_excel_template(self):
        # /import_pack/excel_template sudah ada di controllers/template_download.py
        # Kembalikan action URL agar browser mendownload. 
        return {
            'type': 'ir.actions.act_url',
            'url': '/import_pack/excel_template',
            'target': 'self',
        }

    def download_csv_template(self):
        # /import_pack/csv_template sudah ada di controllers/template_download.py
        return {
            'type': 'ir.actions.act_url',
            'url': '/import_pack/csv_template',
            'target': 'self',
        }

    def action_show_manual_template(self):
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Manual Template'),
                'message': _('Columns: Kode Unit, Deskripsi, Is Pack, Type, Category, Factory Model No, Product Brand, Cal Pack Price, Kode Part, Deskripsi Part, Quantity, UOM, Part Cost'),
                'sticky': False,
            }
        }

    def action_create_sample_file(self):
        sample = "BUNDLE001,Motor Package Set,TRUE,product,All / Saleable,MP-2024-001,Industrial Brand,TRUE,MOTOR001,Motor 1HP,1,Unit,1500000"
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Sample Row'),
                'message': sample,
                'sticky': False,
            }
        }

    # --- Utilitas baca file ---
    def _read_rows(self):
        """Baca rows dari file CSV/XLSX dengan temporary file untuk large files."""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError(_("No file uploaded."))
        
        import tempfile
        import os
        
        rows = []
        ftype = self.import_type
        if not ftype and self.file_name:
            if self.file_name.lower().endswith('.xlsx'):
                ftype = 'xlsx'
            else:
                ftype = 'csv'

        # Decode base64 file
        try:
            content = base64.b64decode(self.file_data)
        except Exception as e:
            raise UserError(_("Failed to decode file: %s") % str(e))

        # ===== XLSX PARSER dengan Temporary File =====
        if ftype == 'xlsx':
            if not openpyxl:
                raise UserError(_("openpyxl is not installed on the server."))
            
            # Buat temporary file
            temp_file = None
            try:
                # Buat temp file dengan suffix .xlsx
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as temp_file:
                    temp_file.write(content)
                    temp_path = temp_file.name
                
                # Baca dengan openpyxl (read_only=True untuk large files)
                wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                ws = wb.active
                
                # Baca header
                header_row = next(ws.iter_rows(min_row=1, max_row=1))
                header = [self.S(c.value) for c in header_row]
                
                # Baca data rows secara streaming
                row_count = 0
                for r in ws.iter_rows(min_row=2, values_only=True):
                    # Skip empty rows
                    if not any(r):
                        continue
                    row = {header[i]: r[i] if i < len(r) else None for i in range(len(header))}
                    rows.append(row)
                    row_count += 1
                    
                    # Log progress setiap 100 rows
                    if row_count % 100 == 0:
                        import logging
                        _logger = logging.getLogger(__name__)
                        _logger.info(f"Parsed {row_count} rows...")
                
                wb.close()
                
            finally:
                # Cleanup temporary file
                if temp_file and os.path.exists(temp_path):
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
        
        # ===== CSV PARSER dengan Temporary File =====
        else:
            temp_file = None
            try:
                # Buat temp file untuk CSV
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.csv', delete=False) as temp_file:
                    temp_file.write(content)
                    temp_path = temp_file.name
                
                # Baca CSV secara streaming
                row_count = 0
                with open(temp_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        # Skip empty rows
                        if not any(row.values()):
                            continue
                        rows.append(row)
                        row_count += 1
                        
                        # Log progress setiap 100 rows
                        if row_count % 100 == 0:
                            import logging
                            _logger = logging.getLogger(__name__)
                            _logger.info(f"Parsed {row_count} rows...")
            
            finally:
                # Cleanup temporary file
                if temp_file and os.path.exists(temp_path):
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
        
        return rows

    # --- Import utama (dipanggil tombol di view) ---
    def button_import(self):
        self.ensure_one()
        
        # ===== VALIDASI AWAL =====
        if not self.file_data:
            raise UserError(_("Please upload a file first."))
        
        # ===== BACA FILE SECARA STREAMING =====
        import logging
        _logger = logging.getLogger(__name__)
        _logger.info("Starting file parsing (streaming mode)...")
        
        try:
            rows = self._read_rows()
        except MemoryError:
            raise UserError(_(
                "File too large to process.\n\n"
                "Solutions:\n"
                "1. Split file into smaller chunks (max 300-500 rows per file)\n"
                "2. Use CSV format instead of Excel\n"
                "3. Contact system administrator to increase server memory"
            ))
        
        if not rows:
            raise UserError(_("No data rows detected in file."))
        
        _logger.info(f"File parsed successfully: {len(rows)} rows found")
        
        # ===== BATCH CONFIGURATION =====
        BATCH_SIZE = 25  # Kurangi ke 20 jika masih error
        
        def G(row, key):
            return self.S(row.get(key))

        Product = self.env['product.product']
        Template = self.env['product.template']
        Pack = self.env['product.pack']
        Uom = self.env['uom.uom']

        uom_cache = {}
        def get_uom_id(name):
            if not name:
                return False
            key = name.lower()
            if key in uom_cache:
                return uom_cache[key]
            rec = Uom.search([('name', 'ilike', name)], limit=1)
            uom_cache[key] = rec.id or False
            return uom_cache[key]

        # Group by bundle
        by_bundle = {}
        for r in rows:
            bundle_code = G(r, 'Kode Unit') or G(r, 'Kode unit') or G(r, 'kode unit')
            if not bundle_code:
                continue  # Skip rows tanpa Kode Unit
            by_bundle.setdefault(bundle_code, []).append(r)

        if not by_bundle:
            raise UserError(_("No valid bundle data found in file."))

        # ===== SPLIT INTO BATCHES =====
        bundle_items = list(by_bundle.items())
        total_bundles = len(bundle_items)
        total_batches = (total_bundles + BATCH_SIZE - 1) // BATCH_SIZE
        
        _logger.info(f"Processing {total_bundles} bundles in {total_batches} batches")
        
        created_tmpl = 0
        updated_tmpl = 0
        created_lines = 0
        updated_lines = 0
        errors = []

        # ===== PROCESS IN BATCHES =====
        for batch_num in range(total_batches):
            start_idx = batch_num * BATCH_SIZE
            end_idx = min(start_idx + BATCH_SIZE, total_bundles)
            batch = bundle_items[start_idx:end_idx]
            
            _logger.info(f"Batch {batch_num + 1}/{total_batches}: Processing bundles {start_idx+1} to {end_idx}")

            try:
                # Process current batch
                for bundle_code, bundle_rows in batch:
                    try:
                        first = bundle_rows[0]
                        bundle_name = G(first, 'Deskripsi') or bundle_code
                        is_pack = self.B(first.get('Is Pack') or True)
                        bundle_type = G(first, 'Type') or 'product'
                        cal_pack_price = self.B(first.get('Cal Pack Price'))
                        
                        parent_mfg_code = G(first, 'Manufacture Code')
                        parent_factory_model = G(first, 'Factory Model No')
                        parent_brand = G(first, 'Product Brand')
                        parent_category = G(first, 'Category')

                        # Find or create template
                        prod = Product.search([('default_code', '=', bundle_code)], limit=1)
                        if prod:
                            tmpl = prod.product_tmpl_id
                            updated_tmpl += 1
                        else:
                            parent_categ_id = False
                            if parent_category:
                                categ = self.env['product.category'].search([('name', '=', parent_category)], limit=1)
                                if not categ:
                                    categ = self.env['product.category'].create({'name': parent_category})
                                parent_categ_id = categ.id
                            
                            tmpl = Template.create({
                                'name': bundle_name,
                                'type': bundle_type if bundle_type in ('product', 'consu', 'service') else 'product',
                                'is_pack': True,
                                'cal_pack_price': cal_pack_price,
                                'manufacture_code': parent_mfg_code or False,
                                'factory_model_no': parent_factory_model or False,
                                'categ_id': parent_categ_id or self.env.ref('product.product_category_all').id,
                            })
                            prod = tmpl.product_variant_id
                            prod.default_code = bundle_code
                            created_tmpl += 1

                        # Update parent attributes
                        if prod:
                            update_vals = {}
                            if parent_mfg_code and not tmpl.manufacture_code:
                                update_vals['manufacture_code'] = parent_mfg_code
                            if parent_factory_model and not tmpl.factory_model_no:
                                update_vals['factory_model_no'] = parent_factory_model
                            if parent_category:
                                categ = self.env['product.category'].search([('name', '=', parent_category)], limit=1)
                                if not categ:
                                    categ = self.env['product.category'].create({'name': parent_category})
                                if categ and tmpl.categ_id != categ:
                                    update_vals['categ_id'] = categ.id
                            if update_vals:
                                tmpl.write(update_vals)

                        tmpl.is_pack = bool(is_pack)
                        if tmpl.cal_pack_price != cal_pack_price:
                            tmpl.cal_pack_price = cal_pack_price

                        component_rows = [r for r in bundle_rows if self.S(r.get('Kode Part'))]

                        if not component_rows:
                            if self.allow_parent_only:
                                continue
                            else:
                                errors.append(f"Bundle {bundle_code}: No components")
                                continue

                        if self.replace_all:
                            tmpl.pack_ids.unlink()

                        existing_map = {}
                        if self.update_existing and not self.replace_all:
                            for pl in tmpl.pack_ids:
                                existing_map[pl.product_id.id] = pl

                        # Process components
                        for r in component_rows:
                            part_code = G(r, 'Kode Part')
                            part_name = G(r, 'Deskripsi Part') or part_code
                            part_cat_name = G(r, 'Part Category')
                            qty = self.F(r.get('Quantity'))
                            uom_name = G(r, 'UOM')
                            cost = self.F(r.get('Part Cost'))

                            if qty <= 0:
                                errors.append(f"Bundle {bundle_code} / Part {part_code}: Invalid quantity")
                                continue

                            comp = Product.search([('default_code', '=', part_code)], limit=1)
                            if not comp:
                                categ_id = False
                                if part_cat_name:
                                    categ = self.env['product.category'].search([('name', '=', part_cat_name)], limit=1)
                                    if not categ:
                                        categ = self.env['product.category'].create({'name': part_cat_name})
                                    categ_id = categ.id

                                comp_tmpl = Template.create({
                                    'name': part_name or part_code,
                                    'type': 'product',
                                    'categ_id': categ_id or self.env.ref('product.product_category_all').id,
                                })

                                comp = comp_tmpl.product_variant_id
                                comp.default_code = part_code
                                if cost > 0:
                                    comp.standard_price = cost
                            else:
                                write_vals = {}
                                if part_cat_name:
                                    categ = self.env['product.category'].search([('name', '=', part_cat_name)], limit=1)
                                    if not categ:
                                        categ = self.env['product.category'].create({'name': part_cat_name})
                                    write_vals['categ_id'] = categ.id
                                if write_vals:
                                    comp.product_tmpl_id.write(write_vals)

                            if self.update_existing and comp.id in existing_map:
                                pl = existing_map[comp.id]
                                vals = {'qty_uom': qty}
                                if cost > 0 and not comp.standard_price:
                                    vals['standard_price'] = cost
                                pl.write(vals)
                                updated_lines += 1
                            else:
                                Pack.create({
                                    'bi_product_template': tmpl.id,
                                    'product_id': comp.id,
                                    'qty_uom': qty,
                                    'default_code': comp.default_code,
                                    'name': comp.name,
                                    'standard_price': comp.standard_price if comp.standard_price else cost,
                                })
                                created_lines += 1

                        tmpl._recompute_pack_price()
                    
                    except Exception as e:
                        errors.append(f"Bundle {bundle_code}: {str(e)}")
                        _logger.error(f"Error processing bundle {bundle_code}: {e}")
                        continue

                # ===== COMMIT BATCH =====
                self.env.cr.commit()
                _logger.info(f"Batch {batch_num + 1}/{total_batches} committed successfully")
                
            except Exception as e:
                _logger.error(f"Fatal error in batch {batch_num + 1}: {e}")
                self.env.cr.rollback()
                raise UserError(_(f"Error in batch {batch_num + 1}: {str(e)}"))

        # ===== RESULT SUMMARY =====
        message = _(
            'Import completed!\n\n'
            'Templates created: %s\n'
            'Templates updated: %s\n'
            'Components created: %s\n'
            'Components updated: %s'
        ) % (created_tmpl, updated_tmpl, created_lines, updated_lines)
        
        if errors:
            message += _('\n\nWarnings (%s):\n%s') % (len(errors), '\n'.join(errors[:10]))
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Product Pack'),
                'message': message,
                'sticky': True,
                'type': 'success' if not errors else 'warning',
            }
        }